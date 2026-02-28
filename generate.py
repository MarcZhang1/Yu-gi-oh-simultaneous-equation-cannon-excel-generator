#!/usr/bin/env python3
"""
Yu-Gi-Oh! Simultaneous Equation Cannon Excel Generator

Finds all positive integer solutions (fusion, xyz) within the given ranges for:

    fusion + xyz      = stars
    fusion + 2 * xyz  = nb_cards

Usage:
    python generate.py <fusion_min> <fusion_max> <xyz_min> <xyz_max>

Example:
    python generate.py 1 5 2 6   -> results/results 2-6 1-5.xlsx
"""

import argparse
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side


def solve(
    fusion_min: int, fusion_max: int, xyz_min: int, xyz_max: int
) -> list[dict]:
    """
    Enumerate every (fusion, xyz) pair in the given ranges and compute
    the corresponding (stars, nb_cards) from the equation system.
    Returns rows sorted by stars descending, then xyz descending.
    """
    solutions = []
    for fusion in range(fusion_min, fusion_max + 1):
        for xyz in range(xyz_min, xyz_max + 1):
            stars = fusion + xyz           # fusion + xyz    = stars
            nb_cards = fusion + 2 * xyz    # fusion + 2*xyz  = nb_cards
            solutions.append(
                {"stars": stars, "nb_cards": nb_cards, "xyz": xyz, "fusion": fusion}
            )

    # Primary sort: stars descending; secondary: xyz descending
    solutions.sort(key=lambda s: (s["stars"], s["xyz"]), reverse=True)
    return solutions


def generate_excel(solutions: list[dict], output_path: str) -> None:
    """Write *solutions* to an Excel file that matches the example format."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # --- Header row ---
    thin = Side(style="thin")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center")
    header_border = Border(left=thin)

    for col_idx, header in enumerate(["stars", "nb cards", "xyz", "fusion"], start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = header_border

    # --- Data rows ---
    current_stars = None
    group_start_row = None

    for row_idx, sol in enumerate(solutions, start=2):
        stars_val = sol["stars"]

        if stars_val != current_stars:
            # Close the previous stars group (merge if it spanned multiple rows)
            if current_stars is not None and row_idx - 1 > group_start_row:
                ws.merge_cells(
                    start_row=group_start_row, start_column=1,
                    end_row=row_idx - 1, end_column=1,
                )

            # Open a new stars group
            current_stars = stars_val
            group_start_row = row_idx
            ws.cell(row=row_idx, column=1, value=stars_val)
        else:
            # Subsequent row in the same group — leave the cell empty
            ws.cell(row=row_idx, column=1, value=None)

        ws.cell(row=row_idx, column=2, value=sol["nb_cards"])
        ws.cell(row=row_idx, column=3, value=sol["xyz"])
        ws.cell(row=row_idx, column=4, value=sol["fusion"])

    # Close the last stars group
    last_row = len(solutions) + 1
    if group_start_row is not None and last_row > group_start_row:
        ws.merge_cells(
            start_row=group_start_row, start_column=1,
            end_row=last_row, end_column=1,
        )

    # --- Save ---
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"Saved: {output_path}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Generate a Simultaneous Equation Cannon spreadsheet.\n\n"
            "Solves the system:\n"
            "  fusion + xyz     = stars\n"
            "  fusion + 2*xyz   = nb_cards\n\n"
            "for all integer (fusion, xyz) pairs within the given ranges."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("xyz_min",    type=int, help="Minimum XYZ monster rank")
    parser.add_argument("xyz_max",    type=int, help="Maximum XYZ monster rank")
    parser.add_argument("fusion_min", type=int, help="Minimum fusion monster level")
    parser.add_argument("fusion_max", type=int, help="Maximum fusion monster level")

    args = parser.parse_args()

    if args.fusion_min > args.fusion_max:
        parser.error("fusion_min must be ≤ fusion_max")
    if args.xyz_min > args.xyz_max:
        parser.error("xyz_min must be ≤ xyz_max")
    if args.fusion_min < 1 or args.xyz_min < 1:
        parser.error("All level/rank values must be ≥ 1")

    solutions = solve(args.fusion_min, args.fusion_max, args.xyz_min, args.xyz_max)

    script_dir = os.path.dirname(os.path.abspath(__file__))
    filename = f"sec xyz{args.xyz_min}-{args.xyz_max} fusion{args.fusion_min}-{args.fusion_max}.xlsx"
    output_path = os.path.join(script_dir, "results", filename)

    generate_excel(solutions, output_path)


if __name__ == "__main__":
    main()
